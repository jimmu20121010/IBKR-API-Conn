#!/usr/bin/env python3
"""
market_analyzer/main.py
========================
全球指數 + 殖利率 + 商品 + 匯率 + 市場情緒層 總覽系統
"""

import asyncio
asyncio.set_event_loop(asyncio.new_event_loop())

import os, sys, sqlite3, logging, time
from datetime import datetime
from pathlib import Path

import pandas as pd
import requests
import yfinance as yf
from dotenv import load_dotenv

try:
    from ib_insync import IB, Index, util
    IB_AVAILABLE = True
except ImportError:
    IB_AVAILABLE = False
    print("[WARN] ib_insync not installed. Falling back to yfinance.")

from email_report import send_email_report, build_html_report
from sentiment    import fetch_all_sentiment
from technical_analysis import get_index_technical_analysis, get_correlation_matrix

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("analyzer.log", encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)

load_dotenv(Path(__file__).parent / ".env")

IB_HOST      = os.getenv("IB_HOST",      "127.0.0.1")
IB_PORT      = int(os.getenv("IB_PORT",  4002))
IB_CLIENT_ID = int(os.getenv("IB_CLIENT_ID", 12))

GMAIL_SENDER       = os.getenv("EMAIL_FROM", "").strip()
GMAIL_APP_PASSWORD = os.getenv("SMTP_PASS",  "").strip()
GMAIL_RECIPIENT    = os.getenv("EMAIL_TO",   "").strip()

BASE_DIR  = Path(__file__).parent
DB_PATH   = BASE_DIR / "market_analysis.db"
EXCEL_DIR = BASE_DIR / "reports"
EXCEL_DIR.mkdir(exist_ok=True)


# ══════════════════════════════════════════════════════════════════════════════
# 定義清單
# ══════════════════════════════════════════════════════════════════════════════

GLOBAL_INDICES = [
    {"key": "twse",     "symbol": "^TWII",     "name": "台股上市指數",   "region": "亞太"},
    {"key": "kospi",    "symbol": "^KS11",     "name": "韓國 KOSPI",     "region": "亞太"},
    {"key": "nikkei",   "symbol": "^N225",     "name": "日經 225",       "region": "亞太"},
    {"key": "hsi",      "symbol": "^HSI",      "name": "香港恒生",       "region": "亞太"},
    {"key": "sse",      "symbol": "000001.SS", "name": "上海綜合",       "region": "亞太"},
    {"key": "ftse",     "symbol": "^FTSE",     "name": "英國 FTSE 100",  "region": "歐洲"},
    {"key": "dax",      "symbol": "^GDAXI",    "name": "德國 DAX",       "region": "歐洲"},
    {"key": "cac",      "symbol": "^FCHI",     "name": "法國 CAC 40",    "region": "歐洲"},
    {"key": "aex",      "symbol": "^AEX",      "name": "荷蘭 AEX",       "region": "歐洲"},
    {"key": "djia",     "symbol": "^DJI",      "name": "道瓊工業",       "region": "美股"},
    {"key": "nasdaq",   "symbol": "^IXIC",     "name": "那斯達克",       "region": "美股"},
    {"key": "sp500",    "symbol": "^GSPC",     "name": "S&P 500",        "region": "美股"},
    {"key": "sox",      "symbol": "^SOX",      "name": "費城半導體 SOX", "region": "美股"},
    {"key": "chuan_hu", "symbol": "2059.TW",   "name": "川湖 2059",      "region": "台股個股"},
]

YIELD_CURVE = [
    {"key": "us3m",  "symbol": "^IRX",  "name": "美國 3M T-Bill", "tenor": "3M"},
    {"key": "us10y", "symbol": "^TNX",  "name": "美國 10Y",       "tenor": "10Y"},
    {"key": "us30y", "symbol": "^TYX",  "name": "美國 30Y",       "tenor": "30Y"},
]

COMMODITIES = [
    {"key": "gold",     "symbol": "GC=F",  "name": "黃金 (Gold)",    "unit": "USD/oz"},
    {"key": "oil_wti",  "symbol": "CL=F",  "name": "WTI 原油",       "unit": "USD/bbl"},
    {"key": "oil_brent","symbol": "BZ=F",  "name": "Brent 原油",     "unit": "USD/bbl"},
    {"key": "copper",   "symbol": "HG=F",  "name": "COMEX 銅",       "unit": "USD/lb"},
    {"key": "silver",   "symbol": "SI=F",  "name": "白銀 (Silver)",  "unit": "USD/oz"},
    {"key": "natgas",   "symbol": "NG=F",  "name": "天然氣",         "unit": "USD/MMBtu"},
]

FX_RATES = [
    {"key": "dxy",    "symbol": "DX-Y.NYB", "name": "美元指數 DXY", "base": "USD"},
    {"key": "eurusd", "symbol": "EURUSD=X",  "name": "EUR/USD",      "base": "EUR"},
    {"key": "usdjpy", "symbol": "JPY=X",     "name": "USD/JPY",      "base": "USD"},
    {"key": "usdcny", "symbol": "CNY=X",     "name": "USD/CNY",      "base": "USD"},
    {"key": "usdtwd", "symbol": "TWD=X",     "name": "USD/TWD",      "base": "USD"},
    {"key": "usdkrw", "symbol": "KRW=X",     "name": "USD/KRW",      "base": "USD"},
    {"key": "gbpusd", "symbol": "GBPUSD=X",  "name": "GBP/USD",      "base": "GBP"},
]


# ══════════════════════════════════════════════════════════════════════════════
# 報價工具
# ══════════════════════════════════════════════════════════════════════════════

def _pct_from_high(current, high52w):
    try:
        if not current or not high52w or high52w == 0: return None
        return round((current - high52w) / high52w * 100, 2)
    except Exception: return None


def get_quote(symbol, name, is_futures=False, calc_52w=True) -> dict:
    try:
        df    = yf.Ticker(symbol).history(period="5d", auto_adjust=False)
        if df.empty: raise ValueError("No data")
        ser   = df["Close"].dropna()
        close = float(ser.iloc[-1])
        prev  = float(ser.iloc[-2]) if len(ser) >= 2 else close
        chg   = round(close - prev, 6)
        pct   = round((chg / prev) * 100, 2) if prev else 0.0
        high52w = pct_high = None
        if calc_52w:
            try:
                info    = yf.Ticker(symbol).info
                high52w = info.get("fiftyTwoWeekHigh")
                if high52w: high52w = round(float(high52w), 6)
            except Exception: pass
            if not high52w:
                try:
                    dh = yf.Ticker(symbol).history(period="1y", auto_adjust=False)
                    if not dh.empty: high52w = round(float(dh["Close"].dropna().max()), 6)
                except Exception: pass
            pct_high = _pct_from_high(close, high52w)
        log.info(f"[{name}] {close:.4f} {chg:+.4f} ({pct:+.2f}%)")
        return {"value": round(close, 6), "change": chg, "change_pct": pct,
                "high52w": high52w, "pct_from_high": pct_high}
    except Exception as e:
        log.warning(f"[{name}({symbol})] Failed: {e}")
        return {"value": None, "change": None, "change_pct": None,
                "high52w": None, "pct_from_high": None}


def fetch_all_indices() -> dict:
    results = {}
    for idx in GLOBAL_INDICES:
        results[idx["key"]] = get_quote(idx["symbol"], idx["name"]); time.sleep(0.3)
    return results


def fetch_yield_curve() -> dict:
    results = {}
    for y in YIELD_CURVE:
        results[y["key"]] = get_quote(y["symbol"], y["name"], calc_52w=False); time.sleep(0.3)
    try:
        v3m = results.get("us3m",  {}).get("value")
        v10 = results.get("us10y", {}).get("value")
        if v3m and v10:
            spread = round(float(v10) - float(v3m), 3)
            results["spread_3m10y"] = {
                "value": spread,
                "inverted": spread < 0,
                "label": "⚠️ 倒掛（衰退信號）" if spread < 0 else "✅ 正常",
            }
            log.info(f"[Yield] 3M={v3m}% 10Y={v10}% Spread={spread:+.3f}%")
    except Exception as e:
        log.warning(f"[Yield] Spread failed: {e}")
        results["spread_3m10y"] = {"value": None, "inverted": None, "label": "N/A"}
    return results


def fetch_commodities() -> dict:
    results = {}
    for c in COMMODITIES:
        results[c["key"]] = get_quote(c["symbol"], c["name"], is_futures=True); time.sleep(0.3)
    return results


def fetch_fx_rates() -> dict:
    results = {}
    for fx in FX_RATES:
        results[fx["key"]] = get_quote(fx["symbol"], fx["name"], calc_52w=False); time.sleep(0.3)
    return results


# ══════════════════════════════════════════════════════════════════════════════
# IB / VIX / Fear & Greed
# ══════════════════════════════════════════════════════════════════════════════

def connect_ib():
    if not IB_AVAILABLE: return None
    ib = IB()
    try:
        ib.connect(IB_HOST, IB_PORT, clientId=IB_CLIENT_ID, timeout=10)
        log.info(f"[IB] Connected to {IB_HOST}:{IB_PORT}")
        return ib
    except Exception as e:
        log.warning(f"[IB] {e}. Fallback to yfinance."); return None


def get_vix(ib=None) -> float:
    if ib:
        try:
            vix_c = Index("VIX", "CBOE", "USD")
            ib.qualifyContracts(vix_c)
            t = ib.reqMktData(vix_c, "", False, False); ib.sleep(2)
            val = t.last or t.close
            if val and float(val) > 0:
                log.info(f"[VIX] IB: {val}"); return round(float(val), 2)
        except Exception: pass
    try:
        df = yf.Ticker("^VIX").history(period="5d", auto_adjust=False)
        if not df.empty:
            val = round(float(df["Close"].dropna().iloc[-1]), 2)
            log.info(f"[VIX] yfinance: {val}"); return val
    except Exception as e: log.warning(f"[VIX] Failed: {e}")
    return float("nan")


def get_fear_greed() -> dict:
    try:
        import fear_greed
        data = fear_greed.get()
        score = int(float(data["score"]))
        label = str(data["rating"]).replace("_"," ").title()
        log.info(f"[F&G] pkg: {score} ({label})")
        return {"score": score, "label": label}
    except Exception: pass
    try:
        url  = "https://production.dataviz.cnn.io/index/fearandgreed/graphdata"
        hdrs = {"User-Agent": "Mozilla/5.0", "Referer": "https://www.cnn.com/markets/fear-and-greed"}
        j    = requests.get(url, headers=hdrs, timeout=10).json()
        score = int(j["fear_and_greed"]["score"])
        label = str(j["fear_and_greed"]["rating"]).replace("_"," ").title()
        log.info(f"[F&G] CNN: {score} ({label})")
        return {"score": score, "label": label}
    except Exception: pass
    return {"score": 50, "label": "Neutral"}


# ══════════════════════════════════════════════════════════════════════════════
# SQLite
# ══════════════════════════════════════════════════════════════════════════════

def init_db(conn):
    conn.execute("""
        CREATE TABLE IF NOT EXISTS macro_summary (
            id                  INTEGER PRIMARY KEY AUTOINCREMENT,
            run_time            TEXT NOT NULL,
            vix                 REAL,
            fear_greed_score    INTEGER,
            fear_greed_label    TEXT,
            yield_spread        REAL,
            yield_inverted      INTEGER,
            -- AAII
            aaii_bullish        REAL,
            aaii_bearish        REAL,
            aaii_spread         REAL,
            aaii_label          TEXT,
            aaii_date           TEXT,
            -- Put/Call Ratio
            pcr_total           REAL,
            pcr_equity          REAL,
            pcr_label           TEXT,
            -- Market Breadth (NYSE)
            breadth_nyse_adv    INTEGER,
            breadth_nyse_dec    INTEGER,
            breadth_nyse_pct    REAL,
            breadth_nyse_label  TEXT,
            breadth_nsdq_adv    INTEGER,
            breadth_nsdq_dec    INTEGER,
            breadth_nsdq_pct    REAL,
            breadth_nsdq_label  TEXT,
            -- Insider
            insider_buy_count   INTEGER,
            insider_sell_count  INTEGER,
            insider_bias        TEXT,
            note                TEXT
        )
    """)
    conn.commit()
    log.info("[DB] SQLite initialized.")


def save_to_db(conn, run_time, vix, fg, yields, sentiment):
    sp = yields.get("spread_3m10y", {})
    aa = sentiment.get("aaii", {})
    pc = sentiment.get("pcr", {})
    bn = sentiment.get("breadth", {}).get("nyse", {})
    bq = sentiment.get("breadth", {}).get("nasdaq", {})
    ins = sentiment.get("insider", {})

    conn.execute("""
        INSERT INTO macro_summary (
            run_time, vix, fear_greed_score, fear_greed_label,
            yield_spread, yield_inverted,
            aaii_bullish, aaii_bearish, aaii_spread, aaii_label, aaii_date,
            pcr_total, pcr_equity, pcr_label,
            breadth_nyse_adv, breadth_nyse_dec, breadth_nyse_pct, breadth_nyse_label,
            breadth_nsdq_adv, breadth_nsdq_dec, breadth_nsdq_pct, breadth_nsdq_label,
            insider_buy_count, insider_sell_count, insider_bias
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        run_time, vix, fg.get("score"), fg.get("label"),
        sp.get("value"), int(bool(sp.get("inverted"))),
        aa.get("bullish"), aa.get("bearish"), aa.get("bull_bear_spread"),
        aa.get("label"), aa.get("survey_date"),
        pc.get("total"), pc.get("equity"), pc.get("label"),
        bn.get("advances"), bn.get("declines"), bn.get("breadth_pct"), bn.get("label"),
        bq.get("advances"), bq.get("declines"), bq.get("breadth_pct"), bq.get("label"),
        ins.get("buy_count"), ins.get("sell_count"), ins.get("net_bias"),
    ))
    conn.commit()
    log.info("[DB] Saved to macro_summary.")


# ══════════════════════════════════════════════════════════════════════════════
# Excel Export
# ══════════════════════════════════════════════════════════════════════════════

def save_to_excel(run_time, vix, fg, indices, yields, commodities, fx_rates, sentiment, ta_results=None, corr=None) -> Path:
    ts   = datetime.strptime(run_time, "%Y-%m-%d %H:%M:%S").strftime("%Y%m%d_%H%M%S")
    path = EXCEL_DIR / f"market_report_{ts}.xlsx"

    def _row(d, key, fields): return [d.get(k, {}).get(f) for k in [key] for f in fields][0] if False else {f: d.get(key, {}).get(f) for f in fields}

    rows_idx  = [{"地區": i["region"], "指數": i["name"], "Symbol": i["symbol"],
                  **{f: indices.get(i["key"],{}).get(f)
                     for f in ["value","change_pct","high52w","pct_from_high"]}}
                 for i in GLOBAL_INDICES]

    sp = yields.get("spread_3m10y", {})
    rows_yld  = [{"期限": y["tenor"], "名稱": y["name"],
                  "殖利率%": yields.get(y["key"],{}).get("value"),
                  "日變動%": yields.get(y["key"],{}).get("change_pct")}
                 for y in YIELD_CURVE]
    rows_yld.append({"期限": "3M-10Y利差", "名稱": "殖利率曲線利差",
                     "殖利率%": sp.get("value"), "日變動%": None})

    rows_comm = [{"商品": c["name"], "單位": c["unit"],
                  **{f: commodities.get(c["key"],{}).get(f)
                     for f in ["value","change_pct","high52w","pct_from_high"]}}
                 for c in COMMODITIES]

    rows_fx   = [{"匯率對": f["name"], "Symbol": f["symbol"],
                  "現值": fx_rates.get(f["key"],{}).get("value"),
                  "日漲跌%": fx_rates.get(f["key"],{}).get("change_pct")}
                 for f in FX_RATES]

    # 情緒層
    aa  = sentiment.get("aaii", {})
    pc  = sentiment.get("pcr",  {})
    bn  = sentiment.get("breadth", {}).get("nyse", {})
    bq  = sentiment.get("breadth", {}).get("nasdaq", {})
    ins = sentiment.get("insider", {})
    rows_sent = [
        {"指標": "AAII Bullish%",           "數值": aa.get("bullish"),          "說明": aa.get("label"),     "更新": aa.get("survey_date")},
        {"指標": "AAII Bearish%",            "數值": aa.get("bearish"),          "說明": "",                  "更新": ""},
        {"指標": "AAII Bull-Bear Spread",    "數值": aa.get("bull_bear_spread"), "說明": "",                  "更新": ""},
        {"指標": "CBOE Total Put/Call",      "數值": pc.get("total"),            "說明": pc.get("label"),     "更新": "今日"},
        {"指標": "CBOE Equity Put/Call",     "數值": pc.get("equity"),           "說明": "",                  "更新": ""},
        {"指標": "NYSE 上漲股數",            "數值": bn.get("advances"),         "說明": bn.get("label"),     "更新": "今日"},
        {"指標": "NYSE 下跌股數",            "數值": bn.get("declines"),         "說明": "",                  "更新": ""},
        {"指標": "NYSE 市場寬度%",           "數值": bn.get("breadth_pct"),      "說明": "",                  "更新": ""},
        {"指標": "Nasdaq 上漲股數",          "數值": bq.get("advances"),         "說明": bq.get("label"),     "更新": "今日"},
        {"指標": "Nasdaq 下跌股數",          "數值": bq.get("declines"),         "說明": "",                  "更新": ""},
        {"指標": "Insider Buy (3天)",        "數值": ins.get("buy_count"),       "說明": ins.get("net_bias"), "更新": "3日"},
        {"指標": "Insider Sell (3天)",       "數值": ins.get("sell_count"),      "說明": "",                  "更新": ""},
    ]

    rows_meta = [{"run_time": run_time, "vix": vix,
                  "fg_score": fg.get("score"), "fg_label": fg.get("label"),
                  "yield_spread": sp.get("value"), "inverted": sp.get("inverted")}]

    from openpyxl.styles import PatternFill, Font

    def _apply_pct_color(ws, col_name):
        hdrs = [c.value for c in ws[1]]
        if col_name not in hdrs: return
        ci = hdrs.index(col_name) + 1
        for row in ws.iter_rows(min_row=2, min_col=ci, max_col=ci):
            for cell in row:
                try:
                    v = float(cell.value)
                    cell.fill = PatternFill("solid", fgColor=("FFCCCC" if v < -10 else ("FFFACC" if v < -3 else "CCFFCC")))
                except (TypeError, ValueError): pass

    def _autofit(ws):
        bold = Font(bold=True)
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(w + 4, 50)
        for cell in ws[1]: cell.font = bold

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame(rows_idx).to_excel(writer,  sheet_name="全球指數",   index=False)
        pd.DataFrame(rows_yld).to_excel(writer,  sheet_name="殖利率曲線", index=False)
        pd.DataFrame(rows_comm).to_excel(writer, sheet_name="商品",       index=False)
        pd.DataFrame(rows_fx).to_excel(writer,   sheet_name="匯率",       index=False)
        pd.DataFrame(rows_sent).to_excel(writer, sheet_name="市場情緒",   index=False)
        pd.DataFrame(rows_meta).to_excel(writer, sheet_name="Run Info",   index=False)
        if ta_results:
            df_ta = pd.DataFrame(ta_results)[[
                "name", "close", "rsi", "macd_signal",
                "ma200_pct", "ma200_label", "ath_pct", "ath_label"
            ]]
            df_ta.to_excel(writer, sheet_name="Index TA", index=False)

        if corr and corr.get("matrix") is not None:
            corr["matrix"].to_excel(writer, sheet_name="Correlation 30D")        
        _apply_pct_color(writer.sheets["全球指數"], "pct_from_high")
        _apply_pct_color(writer.sheets["商品"],     "pct_from_high")
        for ws in writer.sheets.values(): _autofit(ws)

    log.info(f"[Excel] Saved → {path}")
    return path


# ══════════════════════════════════════════════════════════════════════════════
# Console Summary
# ══════════════════════════════════════════════════════════════════════════════

def print_summary(run_time, vix, fg, indices, yields, commodities, fx_rates, sentiment):
    def fv(v, d=2):  return f"{float(v):,.{d}f}" if v is not None else "N/A"
    def fp(v):       return f"{float(v):+.2f}%" if v is not None else "N/A"

    sep = "=" * 100
    sp  = yields.get("spread_3m10y", {})
    print(f"\n{sep}")
    print(f"  GLOBAL MARKET REPORT  {run_time}")
    print(f"  VIX: {fv(vix)} | F&G: {fg.get('score')} ({fg.get('label')}) | "
          f"3M-10Y利差: {fv(sp.get('value'),3)}%  {sp.get('label','')}")
    print(sep)

    # 情緒層
    aa  = sentiment.get("aaii", {})
    pc  = sentiment.get("pcr",  {})
    bn  = sentiment.get("breadth",{}).get("nyse",   {})
    bq  = sentiment.get("breadth",{}).get("nasdaq", {})
    ins = sentiment.get("insider", {})
    print(f"\n  ══ 市場情緒層 ══")
    print(f"  [AAII {aa.get('survey_date','N/A')}] "
          f"Bull={fv(aa.get('bullish'))}% Bear={fv(aa.get('bearish'))}% "
          f"Spread={fp(aa.get('bull_bear_spread'))}  → {aa.get('label','N/A')}")
    print(f"  [Put/Call Ratio] Total={fv(pc.get('total'))} "
          f"Equity={fv(pc.get('equity'))} → {pc.get('label','N/A')}")
    print(f"  [NYSE Breadth]   Adv={bn.get('advances','N/A')} Dec={bn.get('declines','N/A')} "
          f"寬度={fv(bn.get('breadth_pct'))}% → {bn.get('label','N/A')}")
    print(f"  [Nasdaq Breadth] Adv={bq.get('advances','N/A')} Dec={bq.get('declines','N/A')} "
          f"寬度={fv(bq.get('breadth_pct'))}% → {bq.get('label','N/A')}")
    print(f"  [Insider 3D]     Buy={ins.get('buy_count','N/A')} "
          f"Sell={ins.get('sell_count','N/A')} → {ins.get('net_bias','N/A')}")

    # 股票指數
    print(f"\n  ══ 全球股票指數 ══")
    print(f"  {'指數':<22} {'現價':>14} {'日%':>9} {'52W H':>16} {'距高%':>10}")
    print("  " + "-" * 75)
    prev_r = ""
    for idx in GLOBAL_INDICES:
        d = indices.get(idx["key"], {})
        if idx["region"] != prev_r:
            print(f"\n    [{idx['region']}]"); prev_r = idx["region"]
        print(f"  {idx['name']:<22} {fv(d.get('value')):>14} "
              f"{fp(d.get('change_pct')):>9} {fv(d.get('high52w')):>16} "
              f"{fp(d.get('pct_from_high')):>10}")

    # 殖利率
    print(f"\n  ══ 殖利率曲線 ══")
    for y in YIELD_CURVE:
        d = yields.get(y["key"], {})
        print(f"  {y['name']:<20} {fv(d.get('value'),3):>8}% {fp(d.get('change_pct')):>9}")
    print(f"  {'3M-10Y 利差':<20} {fv(sp.get('value'),3):>8}%  {sp.get('label','')}")

    # 商品
    print(f"\n  ══ 商品 ══")
    for c in COMMODITIES:
        d = commodities.get(c["key"], {})
        print(f"  {c['name']:<22} {fv(d.get('value'),4):>12} {fp(d.get('change_pct')):>9} "
              f"{fv(d.get('high52w'),4):>14} {fp(d.get('pct_from_high')):>10}")

    # 匯率
    print(f"\n  ══ 匯率 ══")
    for fx in FX_RATES:
        d = fx_rates.get(fx["key"], {})
        print(f"  {fx['name']:<20} {fv(d.get('value'),4):>12} {fp(d.get('change_pct')):>9}")

    print("\n" + sep + "\n")


# ══════════════════════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════════════════════
def main():
    run_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log.info(f"=== Market Analyzer START {run_time} ===")

    ib = connect_ib()

    log.info("[1/7] Global indices...")
    indices = fetch_all_indices()

    log.info("[2/7] Yield curve...")
    yields = fetch_yield_curve()

    log.info("[3/7] Commodities...")
    commodities = fetch_commodities()

    log.info("[4/7] FX rates...")
    fx_rates = fetch_fx_rates()

    log.info("[5/7] VIX & Fear/Greed...")
    vix = get_vix(ib)
    fg  = get_fear_greed()

    log.info("[6/7] Market Sentiment Layer...")
    sentiment = fetch_all_sentiment()

    # ── 新增：技術分析層 ──────────────────────────────
    log.info("[6.5/7] Index Technical Analysis...")
    ta_results = get_index_technical_analysis()
    corr       = get_correlation_matrix()
    # ─────────────────────────────────────────────────

    log.info("[7/7] Saving results...")
    conn = sqlite3.connect(DB_PATH)
    init_db(conn)
    save_to_db(conn, run_time, vix, fg, yields, sentiment)
    conn.close()

    excel_path = save_to_excel(run_time, vix, fg, indices, yields,
                               commodities, fx_rates, sentiment,
                               ta_results=ta_results, corr=corr)

    if GMAIL_SENDER and GMAIL_APP_PASSWORD and GMAIL_RECIPIENT:
        html_body = build_html_report(
            run_time, vix, fg,
            indices, GLOBAL_INDICES,
            yields, YIELD_CURVE,
            commodities, COMMODITIES,
            fx_rates, FX_RATES,
            sentiment,
            ta_results=ta_results,   # ← 新增
            corr=corr,               # ← 新增
        )
        send_email_report(
            sender=GMAIL_SENDER, app_password=GMAIL_APP_PASSWORD,
            recipient=GMAIL_RECIPIENT,
            subject=f"📊 Global Market Report — {run_time}",
            html_body=html_body, attachment_path=str(excel_path),
        )
    else:
        log.warning("[Email] Credentials not set. Skipping.")

    if ib: ib.disconnect()

    print_summary(run_time, vix, fg, indices, yields, commodities, fx_rates, sentiment)
    print(f"  Excel : {excel_path}\n  DB    : {DB_PATH}\n")


if __name__ == "__main__":
    main()
